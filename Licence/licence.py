import openpyxl
import random
from datetime import datetime
import smtplib
from email.mime.text import MIMEText
import re

# CONFIGURAÇÕES DO EMAIL
SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
EMAIL_ORIGEM = "nexellbr@gmail.com"       # seu email
SENHA = "srle ofqf dybu qzgk"             # senha de app do Gmail
ASSUNTO = "Sua Licença de Ativação"

# Função para validar email
def email_valido(email):
    if not email:
        return False
    email = str(email).strip()
    padrao = r"^[^@\s]+@[^@\s]+\.[^@\s]+$"
    return re.match(padrao, email) is not None

# Função para enviar email
def enviar_email(destinatario, licenca):
    corpo = f"""
    Olá, segue sua licença de ativação:

    LICENÇA: {licenca}

    Atenciosamente,
    Equipe Nexell
    """

    msg = MIMEText(corpo, "plain", "utf-8")
    msg["Subject"] = ASSUNTO
    msg["From"] = EMAIL_ORIGEM
    msg["To"] = destinatario

    try:
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(EMAIL_ORIGEM, SENHA)
            server.sendmail(EMAIL_ORIGEM, destinatario, msg.as_string())
        print(f"✅ Email enviado para {destinatario}")
    except Exception as e:
        print(f"❌ Erro ao enviar email para {destinatario}: {e}")

# ABRE O EXCEL
caminho = r"C:\Users\Yago Bernardes\Documents\Nexell - CEO\Contabilidade\Licence.xlsx"
wb = openpyxl.load_workbook(caminho)
ws = wb.active

# Os dados começam na linha 4 (linhas 1-3 são cabeçalho)
start_row = 4

for row in range(start_row, ws.max_row + 1):
    cliente = ws.cell(row=row, column=1).value   # Coluna A - Cliente
    pagamento = ws.cell(row=row, column=2).value # Coluna B - Pagamento
    plano = ws.cell(row=row, column=3).value     # Coluna C - Plano
    email = ws.cell(row=row, column=6).value     # Coluna F - Email

    if cliente and (not pagamento or str(pagamento).lower() not in ["true", "verdadeiro"]):
        # Marca pagamento como TRUE
        ws.cell(row=row, column=2).value = "VERDADEIRO"

        # Marca data/hora na Coluna D
        agora = datetime.now()
        ws.cell(row=row, column=4).value = agora.strftime("%d/%m/%Y %H:%M")

        # Gera licença
        numero = random.randint(1000, 9999)
        dia = agora.strftime("%d")
        mes = agora.strftime("%b").upper()   # AUG, SEP, etc.
        hora = agora.strftime("%H")

        licenca = f"{numero}-{dia}P{plano}{mes}{hora}"
        ws.cell(row=row, column=5).value = licenca  # Coluna E = Licença

        print(f"Licença gerada para {cliente}: {licenca}")

        # Envia email se válido
        if email_valido(email):
            email_limpo = email.strip()
            print(f"📧 Tentando enviar para: {email_limpo}")
            enviar_email(email_limpo, licenca)
        else:
            print(f"⚠️ Email inválido encontrado na linha {row}: {email}")

# Salva alterações
wb.save(caminho)
wb.close()
print("✅ Processamento concluído!")
